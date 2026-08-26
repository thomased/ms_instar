#!/usr/bin/env python3
"""Build INSTAR.xlsx from INSTAR.csv.

The CSV is canonical: it is generated from the package's R/items.R and is
what the tooling reads and writes. This script produces a spreadsheet
carrying exactly the same rows and columns, formatted so that a human can
actually read and fill it in (wrapped descriptions, sensible column
widths, frozen header, domain groups colour-coded to match the figure).

Regenerate after any change to the framework:
    python3 code/build_instar_xlsx.py
"""
import csv
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
SRC  = ROOT / "data" / "INSTAR.csv"
OUT  = ROOT / "data" / "INSTAR.xlsx"

# Match the figure palette (R/utils.R .palette)
FOUNDATION = "2E5F8E"
WELFARE    = "3F7A3A"
PAPER      = "5A5A5A"

FOUNDATION_DOMAINS = {"Subjects", "Procedures", "Ethics & Compliance"}

def tint(hex_colour: str, factor: float = 0.90) -> str:
    """Lighten a hex colour towards white."""
    r, g, b = (int(hex_colour[i:i + 2], 16) for i in (0, 2, 4))
    r, g, b = (int(c + (255 - c) * factor) for c in (r, g, b))
    return f"{r:02X}{g:02X}{b:02X}"

WIDTHS = {
    "domain": 20, "item": 34, "item_id": 20,
    "description": 78, "lab": 6, "field": 6, "report": 56,
}

def main() -> int:
    if not SRC.exists():
        print(f"missing {SRC}", file=sys.stderr)
        return 1

    rows = list(csv.DictReader(SRC.open(newline="")))
    header = list(rows[0].keys())

    wb = Workbook()
    ws = wb.active
    ws.title = "INSTAR v1.0"

    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Header
    ws.append([h for h in header])
    for i, name in enumerate(header, start=1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF", size=11)
        c.fill = PatternFill("solid", fgColor="333333")
        c.alignment = Alignment(vertical="center", horizontal="left")
        c.border = border
        ws.column_dimensions[get_column_letter(i)].width = WIDTHS.get(name, 18)
    ws.row_dimensions[1].height = 22

    # Body
    for r, row in enumerate(rows, start=2):
        domain = row["domain"]
        if domain == "How to use":
            fill = PatternFill("solid", fgColor="FFF8E1")
        elif domain == "Framework":
            # Reserved, machine-written row: same neutral tint as the
            # paper-details block, not a domain colour.
            fill = PatternFill("solid", fgColor=tint(PAPER, 0.96))
        elif domain == "Paper details":
            fill = PatternFill("solid", fgColor=tint(PAPER, 0.93))
        elif domain in FOUNDATION_DOMAINS:
            fill = PatternFill("solid", fgColor=tint(FOUNDATION))
        else:
            fill = PatternFill("solid", fgColor=tint(WELFARE))

        for i, name in enumerate(header, start=1):
            c = ws.cell(row=r, column=i, value=row[name])
            c.fill = fill
            c.border = border
            if name in ("description", "report"):
                c.alignment = Alignment(wrap_text=True, vertical="top")
            elif name in ("lab", "field"):
                c.alignment = Alignment(horizontal="center", vertical="top")
            else:
                c.alignment = Alignment(vertical="top", wrap_text=True)
            if name == "domain":
                c.font = Font(bold=True, size=10)
            if name == "item_id":
                c.font = Font(size=9, color="777777")
        ws.row_dimensions[r].height = 58

    # Freeze the header, and everything left of `report`, so the column you
    # type into stays beside the description you are answering.
    report_col = get_column_letter(header.index("report") + 1)
    ws.freeze_panes = "D2"

    # item_id is the machine key: needed when the file is read back, but
    # noise for whoever is filling the sheet in. Hide rather than drop.
    id_col = get_column_letter(header.index("item_id") + 1)
    ws.column_dimensions[id_col].hidden = True

    wb.save(OUT)
    print(f"wrote {OUT}  ({len(rows)} rows, freeze at D2, report col {report_col})")
    return 0

if __name__ == "__main__":
    raise SystemExit(main())
